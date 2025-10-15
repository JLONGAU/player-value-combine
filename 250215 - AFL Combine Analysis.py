#%% Index

#####################################
#                                   #
#              Geelong              #
#        AFL Combine Analysis       #
#              Mar 2025             #
#             Matt Birch            #
#                                   #
#####################################

# Set working directory
import os
#working_directory = 'C:\\Users\\birch\\OneDrive\\Desktop\\Projects\\Tabcorp\\'
working_directory = r"C:\Users\john.long\OneDrive - Geelong Football Club\Documents\player-value-combine"
os.chdir(working_directory)

# Import packages
import warnings
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
from sklearn.preprocessing import StandardScaler
import statsmodels.api as sm

# Set configurations
warnings.filterwarnings("ignore")
pd.set_option('display.max_columns', 100)
pd.set_option('display.float_format', lambda x: '%.2f' % x)

# File paths
#file_path = 'C:\\Users\\birch\\OneDrive\\Desktop\\Projects\\Geelong\\Player Value'
file_path = r"C:\Users\john.long\OneDrive - Geelong Football Club\Documents\player-value-combine"
os.chdir(file_path)

#%% Determine position in first 50 games

# Import Champion data
champion_df = pd.read_csv('02. Processed Data\\AFL Player Stats by Round by Season_Joined_POSfixed.csv')[
    ['Player_ID', 'SEASON', 'ROUND', 'POS']].drop_duplicates()

# Sort data by Player_ID, SEASON, and ROUND to ensure chronological order
champion_df = champion_df.sort_values(by=['Player_ID', 'SEASON', 'ROUND'])

# Remap POS values
pos_mapping = {
    'Wing': 'Mid',
    'Mid-Fwd': 'Mid'
}
champion_df['POS'] = champion_df['POS'].replace(pos_mapping)

# Assign a game number to each player's games in chronological order
champion_df['Game_Number'] = champion_df.groupby('Player_ID').cumcount() + 1

# ---- Get total number of games played per player ----
total_games_df = champion_df.groupby('Player_ID').agg(
    Total_Matches=('Game_Number', 'count')
)

# ---- Filter to first 50 games and get most common position ----
first_50_df = champion_df[champion_df['Game_Number'] <= 50]
most_common_pos_df = first_50_df.groupby('Player_ID').agg(
    Most_Common_POS=('POS', lambda x: x.mode().iloc[0] if not x.mode().empty else None)
)

# ---- Merge the two results ----
champion_df = most_common_pos_df.join(total_games_df).reset_index()

# Change column names
champion_df.columns = ['Player_ID', 'POS', 'Total_Matches']


#%% Data Processing

# Import combine data
combine_df = pd.read_excel('01. Raw Data\\Historical Draft Combine Data\\AFL Combine historical- updated Mar 2025.xlsx')

# Import draft data
draft_df = pd.read_csv('02. Processed Data\\picks_with_dob_df.csv')[['Pick Number', 'Club Name', 'Year', 'Draft Type', 'Is_Father_Son', 'Player_ID']].drop_duplicates()

# Drop rows where important columns are NaN
combine_df = combine_df.dropna(subset=['Name'])
draft_df = draft_df.dropna(subset=['Player_ID'])

### Create Player_ID
# Split "Name" column into first and last name
combine_df[['Last_Name', 'First_Name']] = combine_df['Name'].str.split(', ', n=1, expand=True)

# Ensure there are no NaN values in first name by filling missing ones with an empty string
combine_df['First_Name'] = combine_df['First_Name'].fillna('')

# Convert last name to lowercase and remove non-alphabetic characters
combine_df['Last_Name_Clean'] = combine_df['Last_Name'].str.lower().str.replace(r'[^a-z0-9]', '', regex=True)

# Convert DOB and DOT to datetime format, keeping NaT for invalid values
combine_df['DOB'] = pd.to_datetime(combine_df['DOB'], errors='coerce')
combine_df['DOT'] = pd.to_datetime(combine_df['DOT'], errors='coerce')
combine_df['Birth_Year'] = combine_df['DOB'].dt.year.astype('Int64')
combine_df['Year'] = combine_df['DOT'].dt.year.astype('Int64')

# Create Player_ID
combine_df['Player_ID'] = combine_df['First_Name'] + '_' + combine_df['Last_Name_Clean'] + '_' + combine_df['Birth_Year'].astype(str)

# Drop rows with missing draft year before merging to avoid join errors
combine_df = combine_df.dropna(subset=['Year'])

# Ensure draft data has a clean integer Year column for joining
draft_df['Year'] = pd.to_numeric(draft_df['Year'], errors='coerce').astype('Int64')
draft_df = draft_df.dropna(subset=['Year'])

# Join on whether the player got drafted
combine_df = pd.merge(combine_df, 
                      draft_df, 
                      how = 'left',
                      on = ['Player_ID', 'Year'])

# Join on the players position (<50 matches) and total matches
combine_df = pd.merge(combine_df, 
                      champion_df, 
                      how = 'left',
                      on = ['Player_ID'])

combine_df['POS'] = combine_df['POS'].fillna('No matches')
combine_df['Total_Matches'] = combine_df['Total_Matches'].fillna(0)

# If Club Name is not NaN, then player got drafted
combine_df['Is_Drafted'] = combine_df['Club Name'].notna().astype(int)

# Fill father son with zeros
combine_df['Is_Father_Son'] = combine_df['Is_Father_Son'].fillna(0)

# Rename specific metrics
rename_metrics = {
    # Jumping Tests
    "AbsVJ": "Abs VJ",
    "AbsRunVJ_R": "AbsRun VJ Right",
    "AbsRunVJ_L": "AbsRun VJ Left",
    "RunVJ_R": "Run VJ Right",
    "RunVJ_L": "Run VJ Left",

    # Speed and Agility Tests
    "5m":  "5m Sprint",
    "10m": "10m Sprint",
    "20m": "20m Sprint",
    "Agil": "Agility"
    }

combine_df = combine_df.rename(columns=rename_metrics)

test_metrics = {
    # Anthropometric (Body Measurement) Tests
    "Height": {"higher_is_better": True},  # Taller is preferred
    "Mass": {"higher_is_better": True},  # More mass can be beneficial
    "Tri": {"higher_is_better": False},  # Lower skinfolds indicate lower body fat
    "Scap": {"higher_is_better": False},
    "Bi": {"higher_is_better": False},
    "Supra": {"higher_is_better": False},
    "Ab": {"higher_is_better": False},
    "Thi": {"higher_is_better": False},
    "Calf": {"higher_is_better": False},
    "Sum7": {"higher_is_better": False},  # Total sum of skinfolds (lower is better)
    "S&R_Base": {"higher_is_better": True},  # Higher flexibility is better
    "S&R_Standard": {"higher_is_better": True},

    # Physical Measurement Tests
    "Arm": {"higher_is_better": True},  # Longer arm span can be advantageous
    "Hand": {"higher_is_better": True},  # Bigger hands improve ball handling
    "Reach": {"higher_is_better": True},  # Greater reach helps in marking contests

    # Jumping Tests
    "Abs VJ": {"higher_is_better": True},  # Higher vertical jump is better
    "AbsRun VJ Right": {"higher_is_better": True},
    "AbsRun VJ Left": {"higher_is_better": True},
    "VJ": {"higher_is_better": True},
    "Run VJ Right": {"higher_is_better": True},
    "Run VJ Left": {"higher_is_better": True},

    # Speed and Agility Tests
    "5m Sprint": {"higher_is_better": False},  # Faster time (lower) is better
    "10m Sprint": {"higher_is_better": False},
    "20m Sprint": {"higher_is_better": False},
    "Agility": {"higher_is_better": False},  # Lower agility time is better
    "Shuttle": {"higher_is_better": True},  # Higher distance covered in shuttle test is better

    # Endurance and Aerobic Tests
    "Level": {"higher_is_better": True},  # Beep test level - higher is better
    "Dec": {"higher_is_better": False},  # Deceleration should be minimized
    "VO2": {"higher_is_better": True},  # Higher oxygen capacity is better

    # Yo-Yo Intermittent Recovery Test (YYIR2)
    "YYIR2 Level": {"higher_is_better": True},
    "YYIR2 Distance": {"higher_is_better": True},
    "YYIR2 VO2": {"higher_is_better": True},

    # Endurance Running Tests
    "2Kmin": {"higher_is_better": False},  # Lower 2km time is better
    "2Ksec": {"higher_is_better": False},
    "2K": {"higher_is_better": False},
    "3Kmin": {"higher_is_better": False},  # Lower 3km time is better
    "3Ksec": {"higher_is_better": False},
    "3K": {"higher_is_better": False},

    # Repeat Sprint Tests
    "Repeat sprint 1": {"higher_is_better": False},  # Faster time (lower) is better
    "Repeat sprint 2": {"higher_is_better": False},
    "Repeat sprint 3": {"higher_is_better": False},
    "Repeat sprint 4": {"higher_is_better": False},
    "Repeat sprint 5": {"higher_is_better": False},
    "Repeat sprint 6": {"higher_is_better": False},
    "Repeat sprint total": {"higher_is_better": False},

    # Kicking and Ball Skills Tests
    "Kicking Test": {"higher_is_better": True},  # Higher score means better accuracy
    "Kicking Efficiency": {"higher_is_better": True},
    "Clean Hands Test": {"higher_is_better": True},
    "Hands Efficiency": {"higher_is_better": True},
    "Goal Kicking Test": {"higher_is_better": True},
    "Goal Kicking Efficiency": {"higher_is_better": True},

    # Strength Tests
    "2 km TT": {"higher_is_better": False},  # Faster 2km time is better
    "Bench 2RM": {"higher_is_better": True},  # More weight lifted is better
    "Chin 3RM": {"higher_is_better": True},  # More weight lifted is better
    "ITMP Peak": {"higher_is_better": True},  # More peak power is better

}

# Extract test column names from test_metrics dictionary
test_columns = list(test_metrics.keys())

# Replace zero values with NaN only in the relevant test columns
combine_df[test_columns] = combine_df[test_columns].applymap(lambda x: np.nan if isinstance(x, (int, float)) and x == 0 else x)

# Ensure all test columns are numeric before aggregation
for col in test_metrics.keys():
    combine_df[col] = pd.to_numeric(combine_df[col], errors="coerce")

# Define aggregation functions for test columns
aggregation_dict = {
    "Year": "max",  # Most recent year
    #"Age" : "max",
    "Is_Drafted": "max",  # If drafted in any year, retain 1
    "Is_Father_Son": "max"  # If father-son in any year, retain 1
}

# Add aggregation logic for test columns using test_metrics
for col, properties in test_metrics.items():
    if properties["higher_is_better"]:
        aggregation_dict[col] = "max"  # Best (highest) result
    else:
        aggregation_dict[col] = "min"  # Best (lowest) result

# Aggregate data at Player_ID level
combine_df = combine_df.groupby(["Player_ID", "POS", "Total_Matches"]).agg(aggregation_dict).reset_index()

# One hot encode the POS column
pos_encoded = pd.get_dummies(combine_df['POS'], prefix='POS').astype(int)
combine_df = pd.concat([combine_df, pos_encoded], axis=1)

# Draft success criteria
draft_success_df = pd.DataFrame({
    'Year':          [2011, 2012, 2013, 2014, 2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025],
    'Draft_Success_Matches': [ 139,  139,  140,  141,  124,  109,   93,   76,   59,   45,   34,   23,   12,    0,    0]
})

# Merge draft success threshold into combine_df using Year = Season
combine_df = combine_df.merge(draft_success_df,
                              on='Year', 
                              how='left')

# Create the Draft_Success flag (1 if Total_Matches >= Draft Success threshold)
combine_df['Is_Draft_Success_Matches'] = (combine_df['Total_Matches'] >= combine_df['Draft_Success_Matches']).astype(int)

# Calculate the percentage of expected matches played
combine_df['Pct_Expected_Matches'] = combine_df['Total_Matches'] / combine_df['Draft_Success_Matches']

# Drop the YYIR2 distance and V02 columns
combine_df = combine_df.drop(columns=['YYIR2 Distance', 'YYIR2 VO2'])


# Remove 2025 as we dont have AFL draft data for it yet
#combine_df[combine_df['Year'] == 2025].to_excel('02. Processed Data/AFL Combine Analysis_2025.xlsx')
#combine_df = combine_df[combine_df['Year'] != 2025]


def add_engineered_features(df):
    
    #Adds engineered features to the dataset based on AFL combine test results.
    #These features enhance predictive power by combining key performance indicators
    #relevant to elite AFL recruitment.
    
    
    # 1. Speed Index - A weighted average of sprint times, with shorter sprints being more valuable
    #    (important for acceleration and explosive bursts in AFL gameplay)
    #df["Speed_Index"] = (df["5m"] * 0.4) + (df["10m"] * 0.3) + (df["20m"] * 0.3)

    # 2. Explosive Power Score - Measures the player's vertical jump ability relative to sprint speed
    #    (important for contested marks, tackling, and fast directional changes)
    df["Explosive_Power_Score"] = (df["Abs VJ"] + df["Run VJ Right"] + df["Run VJ Left"]) / df["20m Sprint"]

    # 3. Agility-Speed Ratio - How efficiently a player changes direction relative to their sprint speed
    #    (critical for evading defenders and quick lateral movement)
    df["Agility_Speed_Ratio"] = df["Agility"] / df["20m Sprint"]

    # 4. Endurance-Sprint Ratio - Measures sustained speed over a 2km endurance run
    #    (important for midfielders who need both endurance and speed)
    #df["Endurance_Sprint_Ratio"] = df["2K"] / df["20m"]

    # 5. Speed-Endurance Index - Balances speed (sprint times), repeat sprint ability, and endurance
    #    (players who can sustain high speed over endurance tests are highly valued)
    #df["Speed_Endurance_Index"] = (df["20m"] + df["Repeat sprint total"]) / df["YYIR2 Distance"]

    # 6. Strength-to-Weight Ratio - Evaluates upper body strength relative to body mass
    #    (important for tackling, marking contests, and overall physicality)
    #df["Strength_to_Weight"] = (df["Bench 2RM"] + df["Chin 3RM"]) / df["Mass"]

    # 7. Jump Efficiency Score - Measures how efficiently a player jumps relative to their weight
    #    (high jumpers who maintain power-to-weight ratio are more explosive on the field)
    #df["Jump_Efficiency_Score"] = (df["AbsVJ"] + df["RunVJ_R"] + df["RunVJ_L"]) / df["Mass"]

    # 8. Kicking & Hands Efficiency Score - Combines kicking, handball, and goal-kicking efficiency
    #    (technical skill execution under pressure is critical in the AFL)
    #df["Kicking_Hands_Efficiency"] = (df["Kicking Efficiency"] + df["Hands Efficiency"] + df["Goal Kicking Efficiency"]) / 3

    # 9. Speed-Power Composite Score - Balances explosive power (jumping) with speed (sprint times)
    #    (players who can generate high power while maintaining speed are extremely valuable)
    df["Speed_Power_Composite"] = (df["Abs VJ"] * df["Run VJ Right"] * df["Run VJ Left"]) / (
        df["5m Sprint"] * df["10m Sprint"] * df["20m Sprint"]
    )

    # 10. Repeat Sprint Fatigue Index - Measures how much a player's speed declines over repeated sprints
    #     (important for sustaining high-intensity efforts throughout a match)
    df["Repeat_Sprint_Fatigue_Index"] = (df["Repeat sprint 6"] - df["Repeat sprint 1"]) / df["Repeat sprint 1"]

    return df

# Apply the function to the dataset
combine_df = add_engineered_features(combine_df)

# Define engineered metrics
engineered_metrics = {
    "Explosive_Power_Score": {"higher_is_better": True},  
    "Agility_Speed_Ratio":   {"higher_is_better": False},  
    "Speed_Power_Composite": {"higher_is_better": True},  
    "Repeat_Sprint_Fatigue_Index": {"higher_is_better": False}
}

# Add them to the existing test_metrics dictionary
test_metrics.update(engineered_metrics)

#Export data to excel
#combine_df.to_excel('02. Processed Data/AFL Combine Analysis_All_Years.xlsx')

# Identify all numerical columns dynamically (excluding categorical/text fields)
numeric_columns = combine_df.select_dtypes(include=["number"]).columns.tolist()

# Calculate overall percentage of non-NaN values for each numerical column
overall_completeness = combine_df[numeric_columns].notna().mean() * 100

# Calculate percentage of non-NaN values for each numerical column per birth year
yearly_completeness = combine_df.groupby("Year")[numeric_columns].apply(lambda x: x.notna().mean() * 100)

### Plot completeness by year
# Increase the figure height for better visibility of test names
plt.figure(figsize=(15, 12))

# Plot the heatmap
plt.figure(figsize=(15, 8))
sns.heatmap(yearly_completeness.T, cmap="viridis", annot=False, linewidths=0.5)

plt.title("Completeness of Testing Data Over Time by Test Year", fontsize=14)
plt.xlabel("Test Year", fontsize=12)
plt.ylabel("Test Metrics", fontsize=12)
plt.xticks(rotation=45)
plt.yticks(rotation=0)
plt.show()



def calculate_individual_logistic_coefficients(df, test_metrics, valid_tests, target_value, min_sample_size=50):
    """
    Computes logistic regression coefficients for each test metric individually.

    Steps:
    1. Runs a separate logistic regression model for each test.
    2. Drops missing values instead of imputing.
    3. Collects coefficients to compare importance.

    Parameters:
    - df (pd.DataFrame): The original dataset.
    - test_metrics (dict): Dictionary indicating whether "higher is better" for each test.
    - valid_tests (list): List of test columns that passed completeness filtering.

    Returns:
    - pd.DataFrame: Logistic regression coefficients for each test, sorted by importance.
    """
    # df = combine_df
    #df = df_2018_2023
    #target_value = "Pct_Expected_Matches"
    
    results = []
    
    for test in valid_tests:

        # Drop rows where the test is missing (no imputation)
        df_subset = df[[test, target_value]].dropna()

        if df_subset.shape[0] < min_sample_size:  # Skip tests with too little data
            continue

        X = df_subset[[test]]
        y = df_subset[target_value]

        # 🔄 Reverse sign for "lower is better" tests
        if test in test_metrics and not test_metrics[test]["higher_is_better"]:
            X[test] = -X[test]
            
        # Apply StandardScaler (normalize test values)
        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)
        
        # Add intercept term
        X_scaled = sm.add_constant(X_scaled)
        
        # Fit logistic regression using statsmodels
        if target_value == 'Pct_Expected_Matches':
            model = sm.OLS(y, X_scaled) #For regression
            result = model.fit()
        else:      
            model = sm.Logit(y, X_scaled) #For classification
            result = model.fit(disp=0)       
        
        # Extract coefficient & p-value
        coef = result.params[1]
        p_value = result.pvalues[1]
        
        # Fit logistic regression model
        #model = LogisticRegression(max_iter=500)
        #model.fit(X_scaled, y)

        # Store results
        results.append({"Test Metric": test, "Coefficient": coef, "p-value": p_value})
        #results.append({"Test Metric": test, "Coefficient": model.coef_[0][0]})

    # Convert results to DataFrame. If no results were collected (e.g., every
    # metric was filtered out due to insufficient data), return an empty frame
    # with the expected columns so downstream logic can continue gracefully.
    if not results:
        return pd.DataFrame(columns=["Test Metric", "Coefficient", "p-value"])

    # Ensure the expected columns always exist even if some entries are missing
    # so that downstream sorting does not fail with a KeyError when
    # ``Coefficient`` is absent.
    coef_df = pd.DataFrame(results, columns=["Test Metric", "Coefficient", "p-value"])

    if coef_df.empty:
        return coef_df

    coef_df = coef_df.sort_values(by="Coefficient", ascending=False)
    #coef_df = pd.DataFrame(results).sort_values(by="Coefficient", ascending=False)

    return coef_df



def run_logistic_regression(df, target_value, label_override=None, min_sample_size=50):
    
    #df = combine_df
    #df = df_2018_2023
    #target_value = "Pct_Expected_Matches"
    
    # Remove Tests with Low Overall Completeness
    min_completeness_threshold = 30  # Set threshold (e.g., 25% of players must have this test recorded)
    overall_completeness = df.notna().mean() * 100  # % of non-NaN values for each column
    valid_tests = overall_completeness[overall_completeness > min_completeness_threshold].index.tolist()

    valid_tests = [test for test in valid_tests if test not in [
        "Is_Drafted", "Player_ID", "POS", 'POS_Gen Def', 'POS_Gen Fwd',
        'POS_Key Def', 'POS_Key Fwd', 'POS_Mid', 'POS_No matches', 'POS_Ruck',
        "Year", "Is_Father_Son", 'Draft_Success_Matches', 'Is_Draft_Success_Matches',
        'Pct_Expected_Matches', "Total_Matches"]]

    ### Plot completeness by year
    """
    yearly_completeness = df.groupby("Year")[valid_tests].apply(lambda x: x.notna().mean() * 100)
    
    
    # Plot the heatmap
    plt.figure(figsize=(15, 8))
    sns.heatmap(yearly_completeness.T, cmap="viridis", annot=False, linewidths=0.5)

    plt.title(f"Tests with more than {min_completeness_threshold}% of data", fontsize=14)
    plt.xlabel("Test Year", fontsize=12)
    plt.ylabel("Test Metrics", fontsize=12)
    plt.xticks(rotation=45)
    plt.yticks(rotation=0)
    plt.show()
    """

    # Call the function and store the result
    logistic_coefficients = calculate_individual_logistic_coefficients(
        df,
        test_metrics,
        valid_tests,
        target_value,
        min_sample_size=min_sample_size
    )
    
    # Remove statistically not significant coefficients
    logistic_coefficients = logistic_coefficients[logistic_coefficients['p-value'] <= 0.1]
    
    # Year min and max
    max_year = df['Year'].max()
    min_year = df['Year'].min()
    
    # Select top 10 most important test metrics based on correlation
    top_metrics = logistic_coefficients.head(25)

    if top_metrics.empty:
        print(f"No statistically significant test metrics identified for {label_name}. Skipping plots.")
        top_tests = []
    else:
        # Create a horizontal bar chart
        plt.figure(figsize=(10, 6))

        # Bar colors – use dark navy blue like in your original chart
        bar_color = '#102A43'  # Deep navy tone

        bars = plt.barh(top_metrics["Test Metric"], top_metrics["Coefficient"], color=bar_color)

        # Add coefficient labels to the right of each bar
        for bar in bars:
            width = bar.get_width()
            label_x_pos = width + 0.0005  # Slightly offset to the right for clarity
            plt.text(label_x_pos,
                     bar.get_y() + bar.get_height() / 2,
                     f"{width:.3f}",
                     va='center',
                     ha='left',
                     fontsize=9,
                     color='black')

        # Format chart
        plt.xlabel("Logistic Coefficients - (p-value <= 0.1)", fontsize=12)
        plt.ylabel("Test Metric", fontsize=12)
        plt.title(f"Test Relevance - {target_value} - ({min_year}-{max_year})", fontsize=14)
        plt.gca().invert_yaxis()  # Invert y-axis to have highest correlation at the top

        # Show plot
        plt.tight_layout()
        plt.show()

        # List of top test metrics
        top_tests = list(top_metrics['Test Metric'])
    
    ### 🔍 Trellis of Box Plots: Test value distributions split by Is_Drafted ###
    
    label_name = label_override if label_override else target_value
    
    if top_tests:
        num_tests = len(top_tests)
        num_cols = 4
        num_rows = -(-num_tests // num_cols)  # Ceiling division for rows

        fig, axes = plt.subplots(num_rows, num_cols, figsize=(4*num_cols, 4*num_rows), constrained_layout=True)
        axes = axes.flatten()

        for i, test in enumerate(top_tests):
            if test not in df.columns:
                continue

            # Drop rows with NaNs for this test
            plot_data = df[[test, target_value]].dropna()

            # Map 0 → "No", 1 → "Yes" for x-axis labels
            plot_data[target_value] = plot_data[target_value].map({0: "No", 1: "Yes"})

            sns.boxplot(x=target_value, y=test, data=plot_data, ax=axes[i], palette="Set2")

            axes[i].set_title(test, fontsize=10)
            axes[i].set_xlabel("Drafted", fontsize=9)
            axes[i].set_ylabel("Score", fontsize=9)
            axes[i].tick_params(labelsize=8)

            # 🚫 Clip y-axis to 1st and 99th percentiles
            lower, upper = plot_data[test].quantile([0.01, 0.99])
            axes[i].set_ylim(lower, upper)

        # Hide any unused subplots
        for j in range(i+1, len(axes)):
            fig.delaxes(axes[j])

        fig.suptitle(f"Test Distributions - Split by Draft Success ({label_name})", fontsize=16)

        plt.show()

    else:
        print(f"No box plots generated because no top tests were available for {label_name}.")
    
    if pd.notna(min_year) and pd.notna(max_year):
        min_year_str = str(int(min_year))
        max_year_str = str(int(max_year))
        column_string = f"{target_value} - {min_year_str[-2:]} to {max_year_str[-2:]}"
    else:
        column_string = target_value
    
    logistic_coefficients = logistic_coefficients[['Test Metric', 'Coefficient']]
    logistic_coefficients.columns = ['Test Metric', column_string]
    
    return logistic_coefficients, top_tests


def filter_and_clean_data(df, start_year, end_year, completeness_threshold=80):
    """
    Filters and cleans the dataset for a given year range based on completeness threshold.

    Steps:
    1. Filters `df` to only include rows within the given year range.
    2. Removes tests that do not meet the completeness threshold.
    3. Drops rows where selected valid tests contain NaN values.
    4. Returns the cleaned DataFrame.

    Parameters:
    - df (pd.DataFrame): The original dataset.
    - start_year (int): The starting year for filtering.
    - end_year (int): The ending year for filtering.
    - completeness_threshold (int, optional): Minimum % completeness required for a test to be included (default = 80).

    Returns:
    - pd.DataFrame: The filtered and cleaned dataset.
    """

    # Step 1: Filter for the specific year range
    df_filtered = df[(df["Year"] >= start_year) & (df["Year"] <= end_year)].copy()

    # Step 2: Compute overall completeness for each column
    #overall_completeness = df_filtered.notna().mean() * 100  # Convert to percentage

    # Step 3: Identify valid tests that exceed the completeness threshold
    #valid_tests = overall_completeness[overall_completeness > completeness_threshold].index.tolist()

    # Step 4: Drop rows where selected valid test columns contain NaN values
    #df_filtered = df_filtered.dropna(subset=valid_tests)

    # Step 5: Keep only valid test columns
    #df_filtered = df_filtered[valid_tests]

    return df_filtered

# Apply the function to three different date ranges
df_2007_2023 = filter_and_clean_data(combine_df, 2007, 2023)
df_2018_2024 = filter_and_clean_data(combine_df, 2018, 2024)
df_2018_2023 = filter_and_clean_data(combine_df, 2018, 2023)


# Initialize list to collect all logistic coefficient tables
all_coefficient_tables = []

# Track column labels for each run
column_labels = []

##### Using Is_Drafted as the success metric ########
"""
# All positions, all years
log_coef_all_years, _ = run_logistic_regression(combine_df, "Is_Drafted", label_override="All Positions")
all_coefficient_tables.append(log_coef_all_years)
column_labels.append("Is_Drafted - All Years")

# All positions, recent years
log_coef_recent, _ = run_logistic_regression(df_2018_2024, "Is_Drafted", label_override="All Positions")
all_coefficient_tables.append(log_coef_recent)
column_labels.append("Is_Drafted - Recent Years")

# Each position, recent years
for target_value in pos_encoded.columns:
    log_coef, _ = run_logistic_regression(df_2018_2024, target_value, label_override=target_value)
    all_coefficient_tables.append(log_coef)
    column_labels.append(target_value + " - Recent Years")

# Start with the first table
full_results = all_coefficient_tables[0]

# Merge subsequent tables on "Test Metric"
for i in range(1, len(all_coefficient_tables)):
    full_results = pd.merge(
        full_results,
        all_coefficient_tables[i],
        on="Test Metric",
        how="outer"  # Ensures no test metrics are lost
    )

# Fill NaNs with 0 or another marker (e.g., np.nan if you want to preserve missing info)
full_results = full_results.fillna(0)

full_results.to_excel('text.xlsx')
"""


##### Using Pct_Expected_Matches as the success metric ########

# All positions, all years
log_coef_all_years, _ = run_logistic_regression(df_2007_2023, "Pct_Expected_Matches", label_override="All Positions")
all_coefficient_tables.append(log_coef_all_years)
column_labels.append("Pct_Expected_Matches - All Years")

# All positions, recent years
log_coef_recent, _ = run_logistic_regression(df_2018_2023, "Pct_Expected_Matches", label_override="All Positions")
all_coefficient_tables.append(log_coef_recent)
column_labels.append("Pct_Expected_Matches - Recent Years")

# Each position, recent years
for target_value in pos_encoded.columns:
    log_coef, _ = run_logistic_regression(df_2018_2023, target_value, label_override=target_value)
    all_coefficient_tables.append(log_coef)
    column_labels.append(target_value + " - Recent Years")

# Start with the first table
full_results = all_coefficient_tables[0]

# Merge subsequent tables on "Test Metric"
for i in range(1, len(all_coefficient_tables)):
    full_results = pd.merge(
        full_results,
        all_coefficient_tables[i],
        on="Test Metric",
        how="outer"  # Ensures no test metrics are lost
    )

# Fill NaNs with 0 or another marker (e.g., np.nan if you want to preserve missing info)
full_results = full_results.fillna(0)

full_results.to_excel('text.xlsx')




#%% ML Model


# Save data for further analysis
df_2018_2023 = filter_and_clean_data(combine_df, 2018, 2023)

df_2018_2023.to_excel('02. Processed Data/AFL Combine Analysis_2018_2023.xlsx')

### Run this file in the Anaconda command prompt
#cd C:\\Users\\birch\\OneDrive\\Desktop\\Projects\\Geelong\\Player Value\\03. Notebooks
#python "250408 - AFL Combine Analysis_Model_1_Draft.py"

# Draft Prediction with Parallelization


#%% Feature importance for the Draft model

import os
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import joblib
import shap
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

#Set number of columns to view in output
pd.set_option('display.max_columns', 100)
pd.set_option('display.float_format', lambda x: '%.2f' %x)

# Set the file path
#file_path = 'C:\\Users\\birch\\OneDrive\\Desktop\\Projects\\Geelong\\Player Value'
file_path = r'C:\Users\john.long\OneDrive - Geelong Football Club\Documents\player-value-combine'

# Change directory to the file path
os.chdir(file_path)

# Supported model types
supported_model_types = {
    'RandomForestRegressor', 'GradientBoostingRegressor', 'XGBRegressor',
    'LGBMRegressor', 'CatBoostRegressor', 'MLPRegressor'
}

# Load the feature names
columns_all = pd.read_csv('02. Processed Data/draft_model_feature_names.csv')

#position = ['']
#position = '_POS_Ruck'
#position = '_POS_Mid'
#position = '_POS_Gen_Def'
#position = '_POS_Gen_Fwd'
#position = '_POS_Key_Def'
position = '_POS_Key_Fwd'


# Load the feature data frame
draft_model_feature_df = pd.read_csv(f"02. Processed Data/draft_model_feature_df{position}.csv")

#Define all feature names
feature_names = list(columns_all['feature_names']) 

# Define a function to load models
def load_models(folder_path):
    models = {}
    for filename in os.listdir(folder_path):
        if filename.endswith('.pkl'):
            model_key = filename.split('.')[0]
            model_path = os.path.join(folder_path, filename)
            model = joblib.load(model_path)
            model_type = type(model).__name__
            if model_type in supported_model_types:
                models[model_key] = model
                print(f"Loaded model: {model_key} (Type: {model_type})")
            else:
                print(f"Skipping unsupported model: {model_key} (Type: {model_type})")
    return models

# Define a function to get feature importances from a model
def get_feature_importances(model):
    if hasattr(model, 'feature_importances_'):
        return model.feature_importances_
    elif hasattr(model, 'coef_'):
        return model.coef_[0]
    else:
        return np.array([])
    
# Function to calculate SHAP values for models
def calculate_shap_values(model, data, sample_size=100):
    shap_sample_data = data.sample(sample_size, random_state=42)
    explainer = shap.Explainer(model.predict, shap_sample_data)
    shap_values = explainer(shap_sample_data)
    return np.mean(np.abs(shap_values.values), axis=0)

# Load the model feature mapping and filter for Rank == 1
model_features_df = pd.read_csv(f'05. Results/draft_model_results{position}.csv')
model_features_df = model_features_df[model_features_df['Rank'] == 1]
grouped_features = model_features_df.groupby(['model'])

# Create a DataFrame for features with Model, Position, Rating, and Selected Features
features_df = pd.DataFrame([
    {
        'Model': model,
        'Selected Features': feature_names
    }
    for (model), group in grouped_features
])

# Load the models
models = load_models('04. Models/Draft')

# Initialize a DataFrame to store feature importances
feature_importances_df = pd.DataFrame()
failed_models = []

# Iterate through grouped features and calculate importances
for _, row in features_df.iterrows():
    model_name, selected_features = row['Model'], row['Selected Features']
    model_key = f"model_draft_predict{position}"
    if model_key in models:
        model = models[model_key]
        model_type = type(model).__name__
        try:
            filtered_data = draft_model_feature_df[selected_features]
            if hasattr(model, 'feature_importances_'):
                importances = get_feature_importances(model)
            else:
                importances = calculate_shap_values(model, filtered_data)
            temp_df = pd.DataFrame({
                'Feature': selected_features,
                'Importance': importances
            })
            temp_df['Model'] = model_name[0]  # ✅ fixed
            feature_importances_df = pd.concat([feature_importances_df, temp_df], ignore_index=True)
        except Exception as e:
            failed_models.append((model_name, str(e)))
            print(f"Failed for Model: {model_name}, Error: {e}")


# Normalize, rank, and save feature importances to Excel
feature_importances_df['Importance'] = feature_importances_df.groupby(['Model']
)['Importance'].transform(lambda x: (x - x.min()) / (x.max() - x.min()))

# Print summary of failed models
print("\nSummary of failed models:")
for model_name, model_type in failed_models:
    print(f"Model: {model_name}, Type: {model_type}")

# Group feature importances by Model, Position, and Rating
# Plot each model
for (model_name), group in feature_importances_df.groupby(['Model']):
    group = group.sort_values(by='Importance', ascending=False)
    plt.figure(figsize=(10, 6))
    plt.barh(group['Feature'], group['Importance'])
    plt.xlabel('Importance')
    plt.title(f'Feature Importances for {model_name} {position}')
    plt.gca().invert_yaxis()
    plt.tight_layout()
    plt.show()

# Pivot the DataFrame to have "Feature" as rows and "Pos_Rating" as columns with "Importance" as values
feature_importances_df = feature_importances_df.pivot_table(
    index='Feature',
    columns='Model',
    values='Importance',
    aggfunc='mean'
)
feature_importances_df.reset_index(inplace=True)
feature_importances_df.fillna(0, inplace=True)

# Save to Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "feature_importances_df"

for r_idx, row in enumerate(dataframe_to_rows(feature_importances_df, index=False, header=True), 1):
    ws.append(row)
    if r_idx == 1:
        for cell in ws[r_idx]:
            cell.font = Font(bold=True)

ws.freeze_panes = "B2"
for row in ws.iter_rows(min_row=2, min_col=2, max_col=ws.max_column):
    for cell in row:
        cell.number_format = '0%'

for col in range(2, ws.max_column + 1):
    header = ws.cell(row=1, column=col).value
    if header not in ['Average', 'St Dev']:
        rule = ColorScaleRule(start_type="min", start_color="FFFF99",
                              end_type="max", end_color="33FF00")
        ws.conditional_formatting.add(f"{get_column_letter(col)}2:{get_column_letter(col)}{ws.max_row}", rule)

ws.auto_filter.ref = ws.dimensions
ws.column_dimensions['A'].width = 25
for col in range(2, ws.max_column + 1):
    ws.column_dimensions[get_column_letter(col)].width = 20

wb.save(f"05. Results/draft_feature_importances_df{position}.xlsx")




#%% Predict 

import os
import pandas as pd
from sklearn.impute import SimpleImputer
import matplotlib.pyplot as plt
import openpyxl
import joblib
import shap
from openpyxl.styles import Font
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

# Set working directory
#file_path = 'C:\\Users\\birch\\OneDrive\\Desktop\\Projects\\Geelong\\Player Value'
file_path = r'C:\Users\john.long\OneDrive - Geelong Football Club\Documents\player-value-combine'
os.chdir(file_path)

# Display settings
pd.set_option('display.max_columns', 100)
pd.set_option('display.float_format', lambda x: '%.2f' % x)

# Load 2025 input data
#input_file = "02. Processed Data/AFL Combine Analysis_2025.xlsx"
input_file = "02. Processed Data/AFL Combine Analysis_All_Years.xlsx"
ml_data = pd.read_excel(input_file)

# Load the selected feature list used during model training
feature_file = '02. Processed Data/draft_model_feature_names.csv'
columns_all = list(pd.read_csv(feature_file)['feature_names'])

# Rename specific metrics
rename_metrics = {
    # Jumping Tests
    "Abs VJ": "AbsVJ",
    "AbsRun VJ Right": "AbsRunVJ_R",
    "AbsRun VJ Left": "AbsRunVJ_L",
    "Run VJ Right": "RunVJ_R",
    "Run VJ Left": "RunVJ_L",

    # Speed and Agility Tests
    "5m Sprint": "5m",
    "10m Sprint": "10m",
    "20m Sprint": "20m",
    "Agility": "Agil"
}

ml_data = ml_data.rename(columns=rename_metrics)

# Drop players where there are no test results
ml_data = ml_data.dropna(subset=columns_all, how='all')

# Handle missing values using median imputation (same as training)
X_input = ml_data[columns_all]
imputer = SimpleImputer(strategy='median')
X_input_imputed = imputer.fit_transform(X_input)

# Define unique positions
#positions = ["_POS_Gen_Def", "_POS_Gen_Fwd", "_POS_Key_Def", "_POS_Key_Fwd", "_POS_Mid", "_POS_Ruck"]
positions = ['']

for position in positions:

    # Load the trained regression model
    model_path = f'04. Models/Draft/model_draft_predict{position}.pkl'
    model = joblib.load(model_path)
    
    # Make predictions
    ml_data.loc[X_input.index, 'Draft_Pct_Expected_Matches'] = model.predict(X_input_imputed)
    
    # Select specific columns for output
    ml_data = ml_data[['Player_ID', 'Year', 'Draft_Pct_Expected_Matches'] + columns_all]
    
    # Get top 10 based on predicted % matches
    top_players = ml_data.sort_values(by='Draft_Pct_Expected_Matches', ascending=False).head(3)
    
    # Recreate imputed X for just the top 10
    X_top = top_players[columns_all]
    X_top_imputed = imputer.transform(X_top)
    
    # Explain only top 10
    explainer = shap.Explainer(model.predict, X_top_imputed)
    shap_values_top = explainer(X_top_imputed)
    
    # Plot SHAP force plots for top 10 players
    for i in range(len(top_players)):
        player_name = top_players.iloc[i]['Player_ID']
        print(f"🔎 SHAP Force Plot for: {player_name}")
    
        # Create figure and axis manually
        fig, ax = plt.subplots(figsize=(10, 1.2))  # Adjust height as needed
    
        # Plot force with matplotlib target axis
        shap.plots.force(
            shap_values_top[i],
            feature_names=columns_all,
            matplotlib=True,
            show=False  # Avoid double plotting
        )
    
        # Add title
        plt.title(f"SHAP Force Plot – {player_name}")
        plt.tight_layout()
        plt.show()
        

    # Save the results
    output_file = f'05. Results/Undrafted_draft_predict_pct_matches{position}.xlsx'
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        ml_data.to_excel(writer, index=False, sheet_name='2025 Predictions')
    
        workbook = writer.book
        worksheet = writer.sheets['2025 Predictions']
    
        # Freeze top row
        worksheet.freeze_panes = worksheet['A2']
    
        # Bold headers
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
    
        # Set column width for Player_ID
        worksheet.column_dimensions['A'].width = 30
    
        # Format Draft_Pct_Expected_Matches column (assumed to be column C)
        for row in worksheet.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.number_format = '0.00'
    
        # Apply conditional formatting (Green to Red gradient)
        color_scale = ColorScaleRule(
            start_type='min', start_color='FFAA0000',  # Red
            mid_type='percentile', mid_value=50, mid_color='FFFFFF00',  # Yellow
            end_type='max', end_color='FF00AA00'  # Green
        )
        worksheet.conditional_formatting.add('C2:C{}'.format(worksheet.max_row), color_scale)
    
        # Sort by column C descending (we sort in pandas and re-export)
        ml_data.sort_values(by='Draft_Pct_Expected_Matches', ascending=False, inplace=True)
    
        # Re-write sorted data
        ml_data.to_excel(writer, index=False, sheet_name='2025 Predictions')

    print("✅ 2025 Excel output with formatting saved to:", output_file)

